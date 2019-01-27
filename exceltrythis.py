import openpyxl
import csv, codecs
from bs4 import BeautifulSoup
import requests
import json
from time import sleep
import pprint
from openpyxl.chart import (
    Reference, BarChart, Series, ScatterChart
)

# 1) 지난 시간에 작성한 meltop100.csv 파일을 읽어, meltop100.xlsx 로 저장하시오.
#  (단, 랭킹, 좋아요, 좋아요차이 컬럼은 숫자형식으로 저장 할 것!)


fp = codecs.open("meltop100.csv", "r", encoding = "MS949")

reader = csv.reader(fp, delimiter=',', quotechar='"')

book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"

# for j in range (0,5):
#     for i, cells in enumerate(reader):
#         sheet.cell(row = (i+1), column = (j+1)) = cells[j]


for i, cells in enumerate(reader):
    sheet1.cell(row= (i+1), column = 1).value = cells[0]
    sheet1.cell(row= (i+1), column = 2).value = cells[1]
    sheet1.cell(row= (i+1), column = 3).value = cells[2]
    sheet1.cell(row= (i+1), column = 4).value = cells[3]
    sheet1.cell(row= (i+1), column = 5).value = cells[4]



# for i, cells in enumerate(reader):
#     rank = sheet1.cell(row= (i+1), column = 1)
#     if sheet1['A1'] or sheet1['A101']:
#         rank.value = cells[0]
#     else:
#         rank.value = cells[0]
#         rank.number_format 넘버포맷 안 먹힘

            
# 2) 멜론 Top100 곡들의 `앨범 재킷파일`을 다운받아, meltop100.xlsx 파일의 두번째 시트에 랭킹순으로 작성하시오.
# 	(단, 이미지파일의 크기는 축소하여 보기 좋게 작성 할 것!)

sheet2 = book.create_sheet()
sheet2.title = "두번째 시트"


headers = {
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36"
}

url = "https://www.melon.com/chart/index.htm"

html = requests.get(url, headers = headers)
soup = BeautifulSoup(htmlRes.text, 'html.parser')
trs = soup.select('')




# 3) 상위 Top10의 `좋아요 수`는 BarChart로, `좋아요 차이 수`는 ScatterChart로 세번째 시트에 작성하시오.



sheet3 = book.create_sheet()
sheet3.title = "세번째 시트"

data = Reference(sheet1, min_col=4, 
		min_row=2, max_col=4, max_row=11)
categ = Reference(sheet1, min_col=2,
				 min_row=2, max_row=11)

chart = BarChart()
chart.add_data(data=data)
chart.set_categories(categ)

chart.legend = None 
chart.varyColors = True
chart.title = "Top10 좋아요 수"

sheet3.add_chart(chart, "A2")




chart = ScatterChart()
chart.style = 13
chart.x_axis.title = 'likes'
chart.y_axis.title = 'titles'

xvalues = Reference(sheet1, min_col=2,
			 min_row=2, max_row=11)


values = Reference(sheet1, 
            min_col=5, 
            min_row=2, 
            max_row=11)
series = Series(values, xvalues, 
            title = "Top10 좋아요 수")
chart.series.append(series)

sheet3.add_chart(chart, "A18")



book.save("melontop100.xlsx")



